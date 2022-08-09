import sys
import traceback
import io

from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.views import View
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from num2words import num2words

from BI.constants import USER_ROLES
from BI.utilities import execute_read_query
from user_dashboards.models import *
from datetime import datetime
from user_dashboards.constants import *
from user_dashboards.queries import RECEIPTS_MAIN_TABLE, RECEIPTS_TOTAL_AMOUNTS
from user_dashboards.utils import get_receipt_product_data, download_file
from user_profiles.models import MechanicDetail


class ListMechanic(View):
    def __init__(self):
        super(ListMechanic, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(ListMechanic, self).dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):

        all_mechanics = User.objects.filter(user_role=USER_ROLES['MECHANIC'])\
            .values('id', 'first_name', 'last_name', 'full_name', 'rating', 'image')
        mech_ids = [_['id'] for _ in all_mechanics]
        mech_det = MechanicDetail.objects.filter(user_id__in=mech_ids)
        mech_det_dict = {_.user_id: _.expertise for _ in mech_det}
        over_all_rating = 3
        for _ in all_mechanics:
            if _['id'] in mech_det_dict:
                _.update({'expertise': mech_det_dict[_['id']]})
            name = _['full_name'] if _['full_name'] else f"{_['first_name']} {_['last_name']}"
            _.update({'name': name, 'over_all_rating': _['rating'], 'rem_over_all_rating': 5-_['rating']})
        context = {
            'page_headding': 'Find Mechanic',
            'mechanics': list(all_mechanics),
            # 'table_headding_single': 'Receipt',
            # 'table_headding_plural': 'Receipts',
            # 'has_receipts': False,
            # 'column_names': ['ID', 'Customer Name', 'Active Date', 'Total Amount', 'Paid Amount', 'Due Date', 'Status'],
            # 'customers': Customer.objects.filter(created_by=request.user).values('id', 'name', 'reference')
        }
        return render(request, 'main_listing.html', context)
    def post(self, request, *args, **kwargs):
        data = Receipt.objects.filter(created_by=request.user).order_by('expiry_date')
        search_str = request.POST.get('search_query', '')
        if search_str:
            if search_str.isdigit():
                data = data.filter(Q(customer__name__icontains=search_str)|Q(id=search_str))
            else:
                data = data.filter(customer__name__icontains=search_str)
        response_data = []
        total_amounts_data = execute_read_query(RECEIPTS_TOTAL_AMOUNTS.format(','.join([str(_.id) for _ in data])))
        total_amounts_dict = {_[0]:round(int(_[1]), 2) for _ in total_amounts_data}

        for row in data:
            temp_dict = {}
            temp_dict['id'] = row.id
            temp_dict['customer__name'] = row.customer.name
            temp_dict['active_date'] = row.active_date.strftime(GENERAL_DATE)
            temp_dict['paid_amount'] = row.paid_amount
            temp_dict['days_left'] = (row.expiry_date - timezone.now()).days
            temp_dict['expiry_date'] = row.expiry_date.strftime(GENERAL_DATE)
            temp_dict['alarm_color'] = ALARM_COLOR[temp_dict['days_left']]
            temp_dict['status'] = row.status
            temp_dict['total_amount'] = total_amounts_dict.get(row.id, 'No products added')
            response_data.append(temp_dict)
        response = {
            'data': response_data,
            'success': True
        }
        return JsonResponse(data=response)


class Receipts(View):
    def __init__(self):
        super(Receipts, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(Receipts, self).dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        context = {
            'page_headding': 'Receipts',
            'table_headding_single': 'Receipt',
            'table_headding_plural': 'Receipts',
            'has_receipts': False,
            'column_names': ['ID', 'Customer Name', 'Active Date', 'Total Amount', 'Paid Amount', 'Due Date', 'Status'],
            'customers': Customer.objects.filter(created_by=request.user).values('id', 'name', 'reference')
        }
        return render(request, 'receipts.html', context)
    def post(self, request, *args, **kwargs):
        data = Receipt.objects.filter(created_by=request.user).order_by('expiry_date')
        search_str = request.POST.get('search_query', '')
        if search_str:
            if search_str.isdigit():
                data = data.filter(Q(customer__name__icontains=search_str)|Q(id=search_str))
            else:
                data = data.filter(customer__name__icontains=search_str)
        response_data = []
        total_amounts_data = execute_read_query(RECEIPTS_TOTAL_AMOUNTS.format(','.join([str(_.id) for _ in data])))
        total_amounts_dict = {_[0]:round(int(_[1]), 2) for _ in total_amounts_data}

        for row in data:
            temp_dict = {}
            temp_dict['id'] = row.id
            temp_dict['customer__name'] = row.customer.name
            temp_dict['active_date'] = row.active_date.strftime(GENERAL_DATE)
            temp_dict['paid_amount'] = row.paid_amount
            temp_dict['days_left'] = (row.expiry_date - timezone.now()).days
            temp_dict['expiry_date'] = row.expiry_date.strftime(GENERAL_DATE)
            temp_dict['alarm_color'] = ALARM_COLOR[temp_dict['days_left']]
            temp_dict['status'] = row.status
            temp_dict['total_amount'] = total_amounts_dict.get(row.id, 'No products added')
            response_data.append(temp_dict)
        response = {
            'data': response_data,
            'success': True
        }
        return JsonResponse(data=response)


class DeleteReceipt(View):
    def __init__(self):
        super(DeleteReceipt, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(DeleteReceipt, self).dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            receipt_id = int(request.POST.get('receipt_id', 0))
            if not receipt_id:
                raise Exception('No receipt_id given')
            Receipt.objects.filter(id=receipt_id).delete()
            response = {
                'success': True,
                'message': 'Receipt Successfully Deleted'
            }
            return JsonResponse(data=response)
        except Exception as e:
            return JsonResponse(data=self.response_data)


class CreateReceipt(View):
    def __init__(self):
        super(CreateReceipt, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(CreateReceipt, self).dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            customer_id = int(request.POST.get('customer_id', 0))
            due_date = request.POST.get('due_date', '')
            due_date = datetime.strptime(due_date, '%Y-%m-%d')
            Receipt(customer_id=customer_id,
                    expiry_date=due_date,
                    created_by=request.user,
                    status='active'
                    ).save()
            self.response_data = {
                'success': True,
                'message': 'Receipt Created!'
            }
            return JsonResponse(data=self.response_data)
        except Exception as e:
            return JsonResponse(data=self.response_data)


class DeleteCustomer(View):
    def __init__(self):
        super(DeleteCustomer, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(DeleteCustomer, self).dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            customer_id = int(request.POST.get('customer_id', 0))
            if not customer_id:
                raise Exception('No customer_id given')
            Customer.objects.filter(id=customer_id).delete()
            response = {
                'success': True,
                'message': 'Customer Successfully Deleted'
            }
            return JsonResponse(data=response)
        except Exception as e:
            return JsonResponse(data=self.response_data)


class ReceiptDetails(View):
    def __init__(self):
        super(ReceiptDetails, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(ReceiptDetails, self).dispatch(request, *args, **kwargs)

    def get(self, request, receipt_id):
        context = {
            'page_headding': 'Receipt Details - {}'.format(receipt_id),
            'receipt_id': receipt_id,
            'table_headding_single': 'Product',
            'table_headding_plural': 'Products',
            'has_receipts': False,
            'column_names': ['Product ID', 'Product Name', 'Unit', 'Price', 'Quantity', 'Discount', 'Total'],
            'show_print_button': True
        }
        context['products'] = Product.objects.filter(created_by=request.user).values('id', 'name')
        return render(request, 'receipt_details.html', context)

    def post(self, request, *args, **kwargs):
        receipt_id = request.POST.get('receipt_id', 0)
        data = Bill.objects.filter(receipt_id__id=receipt_id)
        search_str = request.POST.get('search_query', '')
        if search_str:
            if search_str.isdigit():
                data = data.filter(Q(product_id__name__icontains=search_str)|Q(id=search_str))
            else:
                data = data.filter(product_id__name__icontains=search_str)
        response_data = get_receipt_product_data(data=data)
        self.response_data = {
            'data': response_data,
            'success': True
        }
        return JsonResponse(data=self.response_data)


class ExportReceipt(View):
    def __init__(self):
        super(ExportReceipt, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(ExportReceipt, self).dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        receipt_id = request.POST.get('receipt_id', 0)
        try:
            receipt_data = get_receipt_product_data(receipt_id)
            customer_data = Receipt.objects.filter(id=receipt_id).first()
            if customer_data:
                customer_data = customer_data.customer

            # wb = load_workbook('user_dashboards/export_receipt_template.xlsx')

            file_contents = download_file('azil-export-excel-files', 'export_receipt_template.xlsx')
            wb = load_workbook(filename=(io.BytesIO(file_contents)), data_only=True)

            ws = wb.active

            starting_row = 17
            col_name = 1
            col_quantity = 6
            col_price = 7
            col_amount = 8

            style_name = ws.cell(row=starting_row, column=col_name)._style
            style_quantity = ws.cell(row=starting_row, column=col_quantity)._style
            style_price = ws.cell(row=starting_row, column=col_price)._style
            style_amount = ws.cell(row=starting_row, column=col_amount)._style


            ws.insert_rows(idx=starting_row, amount=len(receipt_data)-1) # 1 row is already there.

            thin_border = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))

            for r, data in zip(range(starting_row, starting_row + len(receipt_data)), receipt_data):
                ws.merge_cells(f'A{r}:E{r}')
                c = ws.cell(row=r, column=col_name)
                c.value = data['product_name']
                c._style = style_name
                for _ in ws[f'A{r}:E{r}'][0]:
                    _.border = thin_border
                c = ws.cell(row=r, column=col_quantity)
                c.value = data['quantity']
                c._style = style_quantity
                c = ws.cell(row=r, column=col_price)
                c.value = data['price']
                c._style = style_price
                c = ws.cell(row=r, column=col_amount)
                c.value = data['total']
                c._style = style_amount
            total_bill = sum([_['total'] for _ in receipt_data])

            total_row = starting_row + len(receipt_data)
            ws.merge_cells(f'F{total_row}:G{total_row}')
            ws.cell(row=total_row, column=6).value = 'Total Payable:'
            ws.cell(row=total_row, column=8).value = total_bill

            ws.merge_cells(f'A{total_row}:E{total_row}')
            ws.cell(row=total_row, column=1).value = 'Thank you for shoping!'

            for _ in ws[f'A{total_row}:H{total_row}'][0]:
                _.border = thin_border

            if customer_data:
                ws.cell(row=11, column=2).value = customer_data.name
                ws.cell(row=12, column=2).value = customer_data.address
                ws.cell(row=13, column=2).value = customer_data.phone_number
                ws.cell(row=6, column=6).value = customer_data.id

            # setting bill date:
            ws.cell(row=4, column=8).value = datetime.today().strftime(GENERAL_DATE)

            word_amount_row = total_row + 2
            ws.insert_rows(idx=total_row + 1, amount=1)
            ws.cell(row=word_amount_row, column=1).value = 'Amount in words:'
            ws.cell(row=word_amount_row, column=2).value = num2words(total_bill).replace('and ', '').title() + ' Only'

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=BI-Receipt-{receipt_id}.xlsx'
            wb.save(response)
            return response
        except Exception as e:
            exc_type, exc_value, exc_traceback = sys.exc_info()
            print(repr(traceback.format_exception(exc_type, exc_value, exc_traceback)))
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=BI-Receipt-{receipt_id}.xlsx'
            return response


class RemoveProduct(View):
    def __init__(self):
        super(RemoveProduct, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(RemoveProduct, self).dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        receipt_id = request.POST.get('receipt_id', 0)
        product_id = request.POST.get('product_id', 0)
        bill = Bill.objects.filter(receipt_id_id=receipt_id, product_id_id=product_id)
        product_name = bill.first().product_id.name
        bill.delete()

        response = {
            'message': f"{product_name} removed!",
            'success': True
        }
        return JsonResponse(data=response)



class AddProduct(View):
    def __init__(self):
        super(AddProduct, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(AddProduct, self).dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        receipt_id = int(request.POST.get('receipt_id', 0))
        product_id = int(request.POST.get('product_id', 0))
        quantity = int(request.POST.get('quantity', 0))
        bill = Bill(receipt_id_id=receipt_id, product_id_id=product_id, quantity=quantity, added_by=request.user)
        bill.save()

        product_name = bill.product_id.name

        # context = {
        #     'unit': bill.product_id.name,
        #     'price': bill.product_id.price,
        #     'quantity': bill.quantity,
        #     'discount': bill.product_id.discount,
        #     'total': bill.quantity * bill.product_id.price
        # }

        self.response_data = {
            'message': f"{product_name} Added!",
            'success': True,
            # 'data': context
        }
        return JsonResponse(data=self.response_data)


class ListProducts(View):
    def __init__(self):
        super(ListProducts, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(ListProducts, self).dispatch(request, *args, **kwargs)

    def get(self, request):
        context = {
            'page_headding': f'Product List - {request.user.get_user_admin().get_user_full_name()}',
            'table_headding_single': 'Product',
            'table_headding_plural': 'Products',
            'column_names': ['ID', 'Product Name', 'Unit', 'Price', 'Quantity', 'Discount', 'Added By']
        }
        return render(request, 'products.html', context)

    def post(self, request, *args, **kwargs):
        data = Product.objects.filter(created_by=request.user)
        search_str = request.POST.get('search_query', '')
        if search_str:
            if search_str.isdigit():
                data = data.filter(Q(name__icontains=search_str)|Q(id=search_str))
            else:
                data = data.filter(name__icontains=search_str)
        response_data = []
        for row in data:
            temp_dict = {}
            temp_dict['id'] = row.id
            temp_dict['product_name'] = row.name
            temp_dict['unit'] = row.unit
            temp_dict['price'] = row.price
            temp_dict['discount'] = row.discount
            temp_dict['quantity'] = row.quantity
            temp_dict['added_by'] = row.created_by.username
            response_data.append(temp_dict)
        response = {
            'data': response_data,
            'success': True
        }
        return JsonResponse(data=response)

class ListCustomer(View):
    def __init__(self):
        super(ListCustomer, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(ListCustomer, self).dispatch(request, *args, **kwargs)

    def get(self, request):
        context = {
            'page_headding': f'Customer List - {request.user.get_user_admin().get_user_full_name()}',
            'table_headding_single': 'Customer',
            'table_headding_plural': 'Customers',
            'column_names': ['ID', 'Name', 'Reference', 'Phone', 'Address', 'Added By']
        }
        return render(request, 'customer.html', context)

    def post(self, request, *args, **kwargs):
        data = Customer.objects.filter(created_by=request.user)
        search_str = request.POST.get('search_query', '')
        if search_str:
            if search_str.isdigit():
                data = data.filter(id=search_str)
            else:
                data = data.filter(Q(name__icontains=search_str)|Q(address__icontains=search_str))
        response_data = []
        for row in data:
            temp_dict = {}
            temp_dict['id'] = row.id
            temp_dict['customer_name'] = row.name
            temp_dict['reference'] = row.reference
            temp_dict['phone_number'] = row.phone_number
            temp_dict['address'] = row.address
            temp_dict['added_by'] = row.created_by.username
            response_data.append(temp_dict)
        response = {
            'data': response_data,
            'success': True
        }
        return JsonResponse(data=response)


class CreateProducts(View):
    def __init__(self):
        super(CreateProducts, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(CreateProducts, self).dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        name = request.POST.get('name', None)
        price = int(request.POST.get('price', 0))
        unit = request.POST.get('unit', None)
        quantity = int(request.POST.get('quantity', 0))
        discount = int(request.POST.get('discount', 0))

        new_product = Product(name=name, price=price, unit=unit, quantity=quantity, discount=discount, created_by=request.user)
        new_product.save()
        response = {
            'success': True
        }
        return JsonResponse(data=response)


class CreateCustomer(View):
    def __init__(self):
        super(CreateCustomer, self).__init__()
        self.response_data = {'success': False}

    def dispatch(self, request, *args, **kwargs):
        return super(CreateCustomer, self).dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        name = request.POST.get('customer_name', None)
        reference = request.POST.get('reference', 0)
        phone = request.POST.get('phone', None)
        address = request.POST.get('address', None)
        if phone:
            phone = '+92' + phone[1:].replace('-', '')
        new_customer = Customer(name=name, reference=reference, phone_number=phone, address=address, created_by=request.user)
        new_customer.save()
        self.response_data['success'] = True
        self.response_data['message'] = 'Customer Added'

        return JsonResponse(data=self.response_data)